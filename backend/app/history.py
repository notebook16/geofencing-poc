"""Persistent run history (XLSX)."""

from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

HISTORY_HEADERS = ["username", "started_at", "scope", "row_count", "fields", "filename"]


def history_path() -> Path:
    output_dir = Path(os.getenv("OUTPUT_DIR", "")).expanduser()
    if not output_dir or str(output_dir) == ".":
        project_root = Path(__file__).resolve().parents[2]
        output_dir = project_root / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / "run_history.xlsx"


def append_run(
    *,
    username: str,
    started_at: datetime,
    scope: str,
    row_count: int,
    fields: list[str],
    filename: str,
) -> None:
    path = history_path()
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "history"
        ws.append(HISTORY_HEADERS)

    ws.append(
        [
            username,
            started_at.astimezone(timezone.utc).isoformat(),
            scope,
            row_count,
            ", ".join(fields),
            filename,
        ]
    )
    wb.save(path)


def list_runs(limit: int = 100) -> list[dict[str, Any]]:
    path = history_path()
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
        out.append(item)
    out.reverse()
    return out[:limit]
